"""Interview plan and interviewee management API endpoints."""

import csv
import io
import logging
import os
from datetime import datetime

from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from src.models.database import get_db
from src.models.plan import INTERVIEWEE_STATUS_VALUES, PLAN_STATUS_VALUES, Interviewee, InterviewPlan

router = APIRouter()
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Pydantic Models
# ---------------------------------------------------------------------------


class PlanCreateRequest(BaseModel):
    """Request model for creating an interview plan."""

    name: str = Field(..., min_length=1, max_length=255, description="Plan name")
    description: str | None = Field(None, description="Plan description")
    template_id: str = Field(default="quality_survey", description="Template ID")
    start_date: datetime | None = Field(None, description="Start date")
    end_date: datetime | None = Field(None, description="End date")


class PlanUpdateRequest(BaseModel):
    """Request model for updating an interview plan."""

    name: str | None = Field(None, min_length=1, max_length=255, description="Plan name")
    description: str | None = Field(None, description="Plan description")
    template_id: str | None = Field(None, description="Template ID")
    start_date: datetime | None = Field(None, description="Start date")
    end_date: datetime | None = Field(None, description="End date")
    status: str | None = Field(None, description="Plan status")


class IntervieweeCreateRequest(BaseModel):
    """Request model for creating an interviewee."""

    name: str = Field(..., min_length=1, max_length=100, description="Interviewee name")
    user_id: str | None = Field(None, max_length=100, description="User ID")
    phone: str | None = Field(None, max_length=20, description="Phone number")


class PlanResponse(BaseModel):
    """Response model for an interview plan."""

    id: int
    name: str
    description: str | None = None
    template_id: str
    start_date: datetime | None = None
    end_date: datetime | None = None
    status: str
    created_at: datetime
    updated_at: datetime
    interviewees_count: int = 0


class IntervieweeResponse(BaseModel):
    """Response model for an interviewee."""

    id: int
    plan_id: int
    name: str
    user_id: str | None = None
    phone: str | None = None
    status: str
    invited_at: datetime | None = None
    completed_at: datetime | None = None


class ImportResponse(BaseModel):
    """Response model for import operation."""

    imported_count: int
    skipped_count: int
    errors: list[str] = []


# ---------------------------------------------------------------------------
# Auth Helper
# ---------------------------------------------------------------------------


def verify_api_key(x_api_key: str | None = Header(None)) -> str:
    """Verify API key from X-API-Key header."""
    if not x_api_key:
        raise HTTPException(status_code=401, detail="Missing API key")
    expected_key = os.environ.get("INTERNAL_API_KEY")
    if not expected_key:
        logger.warning("INTERNAL_API_KEY not configured - rejecting all requests")
        raise HTTPException(status_code=401, detail="API key not configured")
    if x_api_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    return x_api_key


# ---------------------------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------------------------


def _plan_to_response(plan: InterviewPlan, db: Session) -> PlanResponse:
    """Convert InterviewPlan model to response."""
    interviewees_count = db.query(Interviewee).filter_by(plan_id=plan.id).count()
    return PlanResponse(
        id=plan.id,
        name=plan.name,
        description=plan.description,
        template_id=plan.template_id or "quality_survey",
        start_date=plan.start_date,
        end_date=plan.end_date,
        status=plan.status,
        created_at=plan.created_at,
        updated_at=plan.updated_at,
        interviewees_count=interviewees_count,
    )


def _interviewee_to_response(interviewee: Interviewee) -> IntervieweeResponse:
    """Convert Interviewee model to response."""
    return IntervieweeResponse(
        id=interviewee.id,
        plan_id=interviewee.plan_id,
        name=interviewee.name,
        user_id=interviewee.user_id,
        phone=interviewee.phone,
        status=interviewee.status,
        invited_at=interviewee.invited_at,
        completed_at=interviewee.completed_at,
    )


# ---------------------------------------------------------------------------
# Plan CRUD Endpoints
# ---------------------------------------------------------------------------


@router.post("/plans", response_model=PlanResponse, status_code=201)
def create_plan(
    request: PlanCreateRequest,
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """Create a new interview plan."""
    plan = InterviewPlan(
        name=request.name,
        description=request.description,
        template_id=request.template_id,
        start_date=request.start_date,
        end_date=request.end_date,
        status="draft",
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)
    return _plan_to_response(plan, db)


@router.get("/plans")
def list_plans(
    status: str | None = Query(default=None, description="Filter by status"),
    limit: int = Query(default=20, ge=1, le=100, description="Limit results"),
    offset: int = Query(default=0, ge=0, description="Offset results"),
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """List all interview plans."""
    query = db.query(InterviewPlan)

    if status is not None:
        if status.lower() not in PLAN_STATUS_VALUES:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")
        query = query.filter(InterviewPlan.status == status.lower())

    total = query.count()
    plans = query.offset(offset).limit(limit).all()

    return {
        "plans": [_plan_to_response(p, db) for p in plans],
        "total": total,
    }


@router.get("/plans/{plan_id}", response_model=PlanResponse)
def get_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """Get a specific interview plan."""
    plan = db.query(InterviewPlan).filter_by(id=plan_id).first()
    if plan is None:
        raise HTTPException(status_code=404, detail="Plan not found")
    return _plan_to_response(plan, db)


@router.put("/plans/{plan_id}", response_model=PlanResponse)
def update_plan(
    plan_id: int,
    request: PlanUpdateRequest,
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """Update an interview plan."""
    plan = db.query(InterviewPlan).filter_by(id=plan_id).first()
    if plan is None:
        raise HTTPException(status_code=404, detail="Plan not found")

    # Validate status if provided
    if request.status is not None:
        if request.status.lower() not in PLAN_STATUS_VALUES:
            raise HTTPException(status_code=400, detail=f"Invalid status: {request.status}")
        plan.status = request.status.lower()

    # Update other fields if provided
    if request.name is not None:
        plan.name = request.name
    if request.description is not None:
        plan.description = request.description
    if request.template_id is not None:
        plan.template_id = request.template_id
    if request.start_date is not None:
        plan.start_date = request.start_date
    if request.end_date is not None:
        plan.end_date = request.end_date

    db.commit()
    db.refresh(plan)
    return _plan_to_response(plan, db)


@router.delete("/plans/{plan_id}")
def delete_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """Delete an interview plan and all its interviewees."""
    plan = db.query(InterviewPlan).filter_by(id=plan_id).first()
    if plan is None:
        raise HTTPException(status_code=404, detail="Plan not found")

    db.delete(plan)
    db.commit()
    return {"code": 0, "msg": "success"}


# ---------------------------------------------------------------------------
# Interviewee CRUD Endpoints
# ---------------------------------------------------------------------------


@router.post("/plans/{plan_id}/interviewees", response_model=IntervieweeResponse, status_code=201)
def add_interviewee(
    plan_id: int,
    request: IntervieweeCreateRequest,
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """Add an interviewee to a plan."""
    plan = db.query(InterviewPlan).filter_by(id=plan_id).first()
    if plan is None:
        raise HTTPException(status_code=404, detail="Plan not found")

    interviewee = Interviewee(
        plan_id=plan_id,
        name=request.name,
        user_id=request.user_id,
        phone=request.phone,
        status="pending",
    )
    db.add(interviewee)
    db.commit()
    db.refresh(interviewee)
    return _interviewee_to_response(interviewee)


@router.get("/plans/{plan_id}/interviewees")
def list_interviewees(
    plan_id: int,
    status: str | None = Query(default=None, description="Filter by status"),
    limit: int = Query(default=50, ge=1, le=200, description="Limit results"),
    offset: int = Query(default=0, ge=0, description="Offset results"),
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """List all interviewees for a plan."""
    plan = db.query(InterviewPlan).filter_by(id=plan_id).first()
    if plan is None:
        raise HTTPException(status_code=404, detail="Plan not found")

    query = db.query(Interviewee).filter_by(plan_id=plan_id)

    if status is not None:
        if status.lower() not in INTERVIEWEE_STATUS_VALUES:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")
        query = query.filter(Interviewee.status == status.lower())

    total = query.count()
    interviewees = query.offset(offset).limit(limit).all()

    return {
        "interviewees": [_interviewee_to_response(i) for i in interviewees],
        "total": total,
    }


# ---------------------------------------------------------------------------
# Import Endpoints
# ---------------------------------------------------------------------------


@router.post("/plans/{plan_id}/interviewees/import", response_model=ImportResponse)
def import_interviewees(
    plan_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(verify_api_key),
):
    """Import interviewees from CSV or Excel file."""
    plan = db.query(InterviewPlan).filter_by(id=plan_id).first()
    if plan is None:
        raise HTTPException(status_code=404, detail="Plan not found")

    # Determine file type from content type or filename
    content_type = file.content_type or ""
    filename = file.filename or ""

    imported_count = 0
    skipped_count = 0
    errors: list[str] = []

    # Read file content
    content = file.file.read()

    # Handle CSV files
    if content_type == "text/csv" or filename.endswith(".csv"):
        try:
            # Decode CSV content
            text_content = content.decode("utf-8")
            csv_reader = csv.DictReader(io.StringIO(text_content))

            # Check for required 'name' column
            if csv_reader.fieldnames is None or "name" not in csv_reader.fieldnames:
                raise HTTPException(status_code=400, detail="CSV must have 'name' column")

            for row in csv_reader:
                # Skip empty rows
                if not row.get("name") or row["name"].strip() == "":
                    continue

                name = row.get("name", "").strip()
                user_id = row.get("user_id", "").strip() if row.get("user_id") else None
                phone = row.get("phone", "").strip() if row.get("phone") else None

                # Check for duplicate user_id within the plan
                if user_id:
                    existing = db.query(Interviewee).filter_by(plan_id=plan_id, user_id=user_id).first()
                    if existing:
                        skipped_count += 1
                        continue

                interviewee = Interviewee(
                    plan_id=plan_id,
                    name=name,
                    user_id=user_id,
                    phone=phone,
                    status="pending",
                )
                db.add(interviewee)
                imported_count += 1

            db.commit()

        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="Failed to decode CSV file")

    # Handle Excel files
    elif (
        content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or filename.endswith((".xlsx", ".xls"))
    ):
        try:
            import openpyxl

            # Load workbook
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip().lower() if cell.value else "")

            # Check for required 'name' column
            if "name" not in headers:
                raise HTTPException(status_code=400, detail="Excel must have 'name' column")

            # Get column indices
            name_idx = headers.index("name")
            user_id_idx = headers.index("user_id") if "user_id" in headers else -1
            phone_idx = headers.index("phone") if "phone" in headers else -1

            # Process rows (skip header row)
            for _row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                name_cell = row[name_idx].value
                if not name_cell or str(name_cell).strip() == "":
                    continue

                name = str(name_cell).strip()
                user_id = str(row[user_id_idx].value).strip() if user_id_idx >= 0 and row[user_id_idx].value else None
                phone = str(row[phone_idx].value).strip() if phone_idx >= 0 and row[phone_idx].value else None

                # Check for duplicate user_id within the plan
                if user_id:
                    existing = db.query(Interviewee).filter_by(plan_id=plan_id, user_id=user_id).first()
                    if existing:
                        skipped_count += 1
                        continue

                interviewee = Interviewee(
                    plan_id=plan_id,
                    name=name,
                    user_id=user_id,
                    phone=phone,
                    status="pending",
                )
                db.add(interviewee)
                imported_count += 1

            db.commit()

        except ImportError:
            raise HTTPException(status_code=400, detail="Excel processing requires openpyxl library")
        except Exception as e:
            logger.exception(f"Excel import error: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to process Excel file: {e!s}")

    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Please upload CSV or Excel file.")

    return ImportResponse(
        imported_count=imported_count,
        skipped_count=skipped_count,
        errors=errors,
    )
