"""
Integration tests for interview plan and interviewee API endpoints.

Uses FastAPI TestClient with an in-memory SQLite database so no external
services are required.
"""

import io
import pytest
from datetime import datetime
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from src.api.main import app
from src.models.database import Base, get_db

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

SQLITE_URL = "sqlite:///:memory:"

# Test API key for auth tests
TEST_API_KEY = "test-secret-api-key-12345"


@pytest.fixture(autouse=True)
def setup_test_api_key(monkeypatch):
    """Set up test API key environment variable."""
    monkeypatch.setenv("INTERNAL_API_KEY", TEST_API_KEY)


@pytest.fixture()
def db_engine():
    engine = create_engine(
        SQLITE_URL,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    yield engine
    Base.metadata.drop_all(bind=engine)
    engine.dispose()


@pytest.fixture()
def db_session(db_engine):
    Session = sessionmaker(autocommit=False, autoflush=False, bind=db_engine)
    session = Session()
    try:
        yield session
    finally:
        session.close()


@pytest.fixture()
def client(db_session):
    """TestClient with DB dependency overridden to use in-memory SQLite."""

    def override_get_db():
        yield db_session

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app, headers={"X-API-Key": TEST_API_KEY}) as c:
        yield c
    app.dependency_overrides.clear()


def _make_plan(name="Test Plan", description="Test Description", template_id="quality_survey"):
    """Helper to create a plan request body."""
    return {
        "name": name,
        "description": description,
        "template_id": template_id,
    }


def _make_interviewee(name="Test User", user_id="user_001", phone="13800138000"):
    """Helper to create an interviewee request body."""
    return {
        "name": name,
        "user_id": user_id,
        "phone": phone,
    }


# ---------------------------------------------------------------------------
# POST /api/plans - Create Plan
# ---------------------------------------------------------------------------


class TestCreatePlan:
    """Tests for creating interview plans."""

    def test_create_plan_success(self, client):
        """Test creating a plan successfully."""
        resp = client.post("/api/plans", json=_make_plan())
        assert resp.status_code == 201
        data = resp.json()
        assert data["name"] == "Test Plan"
        assert data["status"] == "draft"
        assert data["id"] is not None

    def test_create_plan_with_dates(self, client):
        """Test creating a plan with start and end dates."""
        plan_data = _make_plan()
        plan_data["start_date"] = "2026-01-01T10:00:00"
        plan_data["end_date"] = "2026-01-31T18:00:00"

        resp = client.post("/api/plans", json=plan_data)
        assert resp.status_code == 201
        data = resp.json()
        assert data["start_date"] == "2026-01-01T10:00:00"
        assert data["end_date"] == "2026-01-31T18:00:00"

    def test_create_plan_missing_name_returns_400(self, client):
        """Test that missing required field 'name' returns 400."""
        resp = client.post("/api/plans", json={"description": "No name"})
        assert resp.status_code == 400

    def test_create_plan_returns_all_fields(self, client):
        """Test that create response includes all required fields."""
        resp = client.post("/api/plans", json=_make_plan())
        data = resp.json()
        for field in ("id", "name", "description", "template_id", "status", "created_at", "updated_at"):
            assert field in data, f"Missing field: {field}"

    def test_create_plan_without_auth_returns_401(self, db_session):
        """Test that missing API key returns 401."""

        def override_get_db():
            yield db_session

        app.dependency_overrides[get_db] = override_get_db
        with TestClient(app) as c:  # No X-API-Key header
            resp = c.post("/api/plans", json=_make_plan())
            assert resp.status_code == 401
        app.dependency_overrides.clear()

    def test_create_plan_invalid_api_key_returns_403(self, db_session):
        """Test that invalid API key returns 403."""

        def override_get_db():
            yield db_session

        app.dependency_overrides[get_db] = override_get_db
        with TestClient(app, headers={"X-API-Key": "wrong-key"}) as c:
            resp = c.post("/api/plans", json=_make_plan())
            assert resp.status_code == 403
        app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# GET /api/plans - List Plans
# ---------------------------------------------------------------------------


class TestListPlans:
    """Tests for listing interview plans."""

    def test_empty_db_returns_empty_list(self, client):
        """Test that empty database returns empty list."""
        resp = client.get("/api/plans")
        assert resp.status_code == 200
        data = resp.json()
        assert data["plans"] == []
        assert data["total"] == 0

    def test_returns_all_plans(self, client):
        """Test that all plans are returned."""
        client.post("/api/plans", json=_make_plan("Plan 1"))
        client.post("/api/plans", json=_make_plan("Plan 2"))

        resp = client.get("/api/plans")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 2
        assert len(data["plans"]) == 2

    def test_filter_by_status(self, client):
        """Test filtering plans by status."""
        client.post("/api/plans", json=_make_plan("Draft Plan"))

        # Create and activate another plan
        resp = client.post("/api/plans", json=_make_plan("Active Plan"))
        plan_id = resp.json()["id"]
        client.put(f"/api/plans/{plan_id}", json={"status": "active"})

        resp = client.get("/api/plans?status=draft")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 1
        assert data["plans"][0]["name"] == "Draft Plan"

    def test_pagination(self, client):
        """Test pagination of plans."""
        for i in range(5):
            client.post("/api/plans", json=_make_plan(f"Plan {i}"))

        resp = client.get("/api/plans?limit=2&offset=1")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 5
        assert len(data["plans"]) == 2

    def test_invalid_status_returns_400(self, client):
        """Test that invalid status returns 400."""
        resp = client.get("/api/plans?status=invalid_status")
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
# GET /api/plans/{id} - Get Plan
# ---------------------------------------------------------------------------


class TestGetPlan:
    """Tests for getting a single plan."""

    def test_get_plan_success(self, client):
        """Test getting a plan successfully."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.get(f"/api/plans/{plan_id}")
        assert resp.status_code == 200
        data = resp.json()
        assert data["name"] == "Test Plan"

    def test_get_plan_not_found_returns_404(self, client):
        """Test that non-existent plan returns 404."""
        resp = client.get("/api/plans/99999")
        assert resp.status_code == 404

    def test_get_plan_with_interviewees(self, client):
        """Test getting a plan includes interviewees count."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        # Add interviewees
        client.post(f"/api/plans/{plan_id}/interviewees", json=_make_interviewee())

        resp = client.get(f"/api/plans/{plan_id}")
        assert resp.status_code == 200
        data = resp.json()
        assert data["interviewees_count"] == 1


# ---------------------------------------------------------------------------
# PUT /api/plans/{id} - Update Plan
# ---------------------------------------------------------------------------


class TestUpdatePlan:
    """Tests for updating a plan."""

    def test_update_plan_name(self, client):
        """Test updating plan name."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.put(f"/api/plans/{plan_id}", json={"name": "Updated Name"})
        assert resp.status_code == 200
        data = resp.json()
        assert data["name"] == "Updated Name"

    def test_update_plan_status(self, client):
        """Test updating plan status."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.put(f"/api/plans/{plan_id}", json={"status": "active"})
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "active"

    def test_update_plan_not_found_returns_404(self, client):
        """Test that updating non-existent plan returns 404."""
        resp = client.put("/api/plans/99999", json={"name": "New Name"})
        assert resp.status_code == 404

    def test_update_plan_invalid_status_returns_400(self, client):
        """Test that invalid status returns 400."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.put(f"/api/plans/{plan_id}", json={"status": "invalid"})
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
# DELETE /api/plans/{id} - Delete Plan
# ---------------------------------------------------------------------------


class TestDeletePlan:
    """Tests for deleting a plan."""

    def test_delete_plan_success(self, client):
        """Test deleting a plan successfully."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.delete(f"/api/plans/{plan_id}")
        assert resp.status_code == 200
        assert resp.json()["code"] == 0

        # Verify it's deleted
        resp = client.get(f"/api/plans/{plan_id}")
        assert resp.status_code == 404

    def test_delete_plan_not_found_returns_404(self, client):
        """Test that deleting non-existent plan returns 404."""
        resp = client.delete("/api/plans/99999")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/plans/{id}/interviewees - Add Interviewee
# ---------------------------------------------------------------------------


class TestAddInterviewee:
    """Tests for adding interviewees to a plan."""

    def test_add_interviewee_success(self, client):
        """Test adding an interviewee successfully."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.post(f"/api/plans/{plan_id}/interviewees", json=_make_interviewee())
        assert resp.status_code == 201
        data = resp.json()
        assert data["name"] == "Test User"
        assert data["plan_id"] == plan_id
        assert data["status"] == "pending"

    def test_add_interviewee_plan_not_found_returns_404(self, client):
        """Test that adding to non-existent plan returns 404."""
        resp = client.post("/api/plans/99999/interviewees", json=_make_interviewee())
        assert resp.status_code == 404

    def test_add_interviewee_missing_name_returns_400(self, client):
        """Test that missing required field returns 400."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.post(f"/api/plans/{plan_id}/interviewees", json={"user_id": "user_001"})
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
# GET /api/plans/{id}/interviewees - List Interviewees
# ---------------------------------------------------------------------------


class TestListInterviewees:
    """Tests for listing interviewees of a plan."""

    def test_list_interviewees_empty(self, client):
        """Test listing interviewees when none exist."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        resp = client.get(f"/api/plans/{plan_id}/interviewees")
        assert resp.status_code == 200
        data = resp.json()
        assert data["interviewees"] == []
        assert data["total"] == 0

    def test_list_interviewees_returns_all(self, client):
        """Test that all interviewees are returned."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        client.post(f"/api/plans/{plan_id}/interviewees", json=_make_interviewee("User 1"))
        client.post(f"/api/plans/{plan_id}/interviewees", json=_make_interviewee("User 2"))

        resp = client.get(f"/api/plans/{plan_id}/interviewees")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 2
        assert len(data["interviewees"]) == 2

    def test_list_interviewees_filter_by_status(self, client):
        """Test filtering interviewees by status."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        client.post(f"/api/plans/{plan_id}/interviewees", json=_make_interviewee("Pending User"))

        resp = client.get(f"/api/plans/{plan_id}/interviewees?status=pending")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 1

    def test_list_interviewees_plan_not_found_returns_404(self, client):
        """Test that listing from non-existent plan returns 404."""
        resp = client.get("/api/plans/99999/interviewees")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/plans/{id}/interviewees/import - Import CSV/Excel
# ---------------------------------------------------------------------------


class TestImportInterviewees:
    """Tests for importing interviewees from CSV/Excel."""

    def test_import_csv_success(self, client):
        """Test importing interviewees from CSV file."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        # Create CSV content
        csv_content = "name,user_id,phone\nJohn Doe,user_001,13800138000\nJane Doe,user_002,13900139000"
        csv_file = io.BytesIO(csv_content.encode("utf-8"))

        resp = client.post(
            f"/api/plans/{plan_id}/interviewees/import",
            files={"file": ("interviewees.csv", csv_file, "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["imported_count"] == 2
        assert data["skipped_count"] == 0

        # Verify interviewees were created
        resp = client.get(f"/api/plans/{plan_id}/interviewees")
        assert resp.json()["total"] == 2

    def test_import_csv_with_empty_lines(self, client):
        """Test importing CSV with empty lines."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        csv_content = "name,user_id,phone\n\nJohn Doe,user_001,13800138000\n\nJane Doe,user_002,13900139000\n"
        csv_file = io.BytesIO(csv_content.encode("utf-8"))

        resp = client.post(
            f"/api/plans/{plan_id}/interviewees/import",
            files={"file": ("interviewees.csv", csv_file, "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["imported_count"] == 2

    def test_import_csv_missing_name_column(self, client):
        """Test that CSV missing 'name' column returns 400."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        csv_content = "user_id,phone\nuser_001,13800138000"
        csv_file = io.BytesIO(csv_content.encode("utf-8"))

        resp = client.post(
            f"/api/plans/{plan_id}/interviewees/import",
            files={"file": ("interviewees.csv", csv_file, "text/csv")},
        )
        assert resp.status_code == 400

    def test_import_excel_success(self, client):
        """Test importing interviewees from Excel file."""
        # Note: This test requires openpyxl to be installed
        # For now, we test that the endpoint exists and handles the file type
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        # Create a simple Excel file (xlsx) using openpyxl if available
        # If not available, we'll test with an empty file to verify endpoint exists
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "name"
            ws["B1"] = "user_id"
            ws["C1"] = "phone"
            ws["A2"] = "Excel User"
            ws["B2"] = "excel_001"
            ws["C2"] = "13800138000"

            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            resp = client.post(
                f"/api/plans/{plan_id}/interviewees/import",
                files={
                    "file": (
                        "interviewees.xlsx",
                        excel_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert resp.status_code == 200
            data = resp.json()
            assert data["imported_count"] == 1
        except ImportError:
            # openpyxl not available, skip the test with a warning
            pytest.skip("openpyxl not installed, skipping Excel import test")

    def test_import_plan_not_found_returns_404(self, client):
        """Test that importing to non-existent plan returns 404."""
        csv_content = "name,user_id,phone\nJohn Doe,user_001,13800138000"
        csv_file = io.BytesIO(csv_content.encode("utf-8"))

        resp = client.post(
            "/api/plans/99999/interviewees/import",
            files={"file": ("interviewees.csv", csv_file, "text/csv")},
        )
        assert resp.status_code == 404

    def test_import_invalid_file_type_returns_400(self, client):
        """Test that invalid file type returns 400."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        # Create a text file (not CSV or Excel)
        text_file = io.BytesIO(b"This is not a valid CSV or Excel file")

        resp = client.post(
            f"/api/plans/{plan_id}/interviewees/import",
            files={"file": ("interviewees.txt", text_file, "text/plain")},
        )
        assert resp.status_code == 400

    def test_import_duplicate_user_id(self, client):
        """Test importing with duplicate user_id (should skip duplicates)."""
        resp = client.post("/api/plans", json=_make_plan())
        plan_id = resp.json()["id"]

        # First import
        csv_content = "name,user_id,phone\nJohn Doe,user_001,13800138000"
        csv_file = io.BytesIO(csv_content.encode("utf-8"))
        client.post(
            f"/api/plans/{plan_id}/interviewees/import",
            files={"file": ("interviewees.csv", csv_file, "text/csv")},
        )

        # Second import with same user_id
        csv_file2 = io.BytesIO(csv_content.encode("utf-8"))
        resp = client.post(
            f"/api/plans/{plan_id}/interviewees/import",
            files={"file": ("interviewees2.csv", csv_file2, "text/csv")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["imported_count"] == 0
        assert data["skipped_count"] == 1
